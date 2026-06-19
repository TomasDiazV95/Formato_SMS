
from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import datetime
from typing import Optional
from difflib import SequenceMatcher
import re
import unicodedata

import pandas as pd
from openpyxl import load_workbook

from repositories import ejecutivos_repo
from services import config_store
from services.contact_dedupe import dedupe_by_column_keep_first, dedupe_by_column_keep_first_normalized
from utils.paths import archive_path, config_path, PROJECT_ROOT

TEMPLATE_COLUMNS_TANNER = [
    "INSTITUCIÓN",
    "SEGMENTOINSTITUCIÓN",
    "message_id",
    "RUT+DV",
    "OPERACION",
    "dest_email",
    "name_from",
    "mail_from",
    "NOMBRE_AGENTE",
    "MAIL_AGENTE",
    "PHONO_AGENTE",
    "MES",
    "ANO",
]

TEMPLATE_COLUMNS_SCJ = [
    "INSTITUCIÓN",
    "SEGMENTOINSTITUCIÓN",
    "message_id",
    "PLANTILLA",
    "RUT",
    "NUM_OP",
    "name_from",
    "dest_email",
    "mail_from",
    "MAIL_AGENTE",
    "PHONO_AGENTE",
    "telefono",
    "NOMBRE_AGENTE",
]

TEMPLATE_COLUMNS_SC_TELEFONIA = [
    "INSTITUCIÓN",
    "SEGMENTOINSTITUCIÓN",
    "message_id",
    "PLANTILLA",
    "CLIENTE",
    "NRO_OPERACION",
    "dest_email",
    "name_from",
    "mail_from",
    "CORREO",
]

TEMPLATE_COLUMNS_SC_TELEFONIA_MEDIOS_PAGO = [
    "INSTITUCIÓN",
    "SEGMENTOINSTITUCIÓN",
    "message_id",
    "PLANTILLA",
    "RUT",
    "dest_email",
    "name_from",
    "mail_from",
    "CORREO",
]

TEMPLATE_COLUMNS_ITAU_VENCIDA = [
    "INSTITUCIÓN",
    "SEGMENTOINSTITUCIÓN",
    "message_id",
    "RUTDV",
    "RUT ",
    "DV",
    "OPERACIONES ",
    "NOMBRE ",
    "APELLIDO_1",
    "APELLIDO_2",
    "NOMBRE COMPLETO",
    "dest_email",
    "name_from",
    "mail_from",
    "CORREO",
    "EJECUTIVO",
    "SUPERVISOR",
    "MAILS SUPERVISOR",
    "TELEFONO SUPERVISOR",
    "CARTERA",
    "CORREO_RECEPCION",
    "FONO",
    "MES_CURSO",
    "ANO_CURSO",
    "DIA_RENE",
    "MES_RENE",
    "ANO_RENE",
    "MARCA WEB Y SUCURSAL",
]

TEMPLATE_COLUMNS_ITAU_CASTIGO = [
    "INSTITUCIÓN",
    "SEGMENTOINSTITUCIÓN",
    "message_id",
    "PLANTILLA",
    "RUT",
    "OPERACION",
    "dest_email",
    "name_from",
    "mail_from",
    "CORREO",
]

TEMPLATE_COLUMNS_BIT = [
    "INSTITUCIÓN",
    "SEGMENTOINSTITUCIÓN",
    "message_id",
    "RUT",
    "OPERACION",
    "CLIENTE",
    "dest_email",
    "name_from",
    "mail_from",
    "MAIL_AGENTE",
]

SPANISH_MONTHS = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
]

AGENTE_COLUMN_ALIASES = {
    "nombre_agente",
    "agente",
    "agentes",
    "ejecutivo",
    "ejecutivos",
    "carterizado",
    "carterizados",
    "carterizado abril",
    "nombre agente",
}

RUT_COLUMN_ALIASES = {
    "rut",
    "rutdv",
    "rut+dv",
    "rut-dv",
    "rut dv",
    "rut_dv",
    "rut cliente",
    "rut_cliente",
    "id_cliente",
}

OPERATION_COLUMN_ALIASES = {
    "operacion",
    "operación",
    "ope",
    "op",
    "nro_operacion",
    "nro operación",
    "nro_operación",
    "nro ope",
    "nro_ope",
    "n operación",
    "n_operacion",
    "n_operación",
    "numero_operacion",
    "numero operación",
    "número operación",
    "número_operación",
    "num_op",
    "nro documento",
    "nro_documento",
    "id_credito",
    "contrato",
}

EMAIL_COLUMN_ALIASES = {
    "dest_email",
    "email",
    "e-mail",
    "correo",
    "correo electronico",
    "correo electrónico",
    "mail",
    "dest_mail",
    "email_cliente",
    "mail_cliente",
}

TANNER_REQUIRED_COLUMN_LABELS = {
    "RUT+DV": RUT_COLUMN_ALIASES,
    "OPERACION": OPERATION_COLUMN_ALIASES,
    "dest_email": EMAIL_COLUMN_ALIASES,
    "NOMBRE_AGENTE": AGENTE_COLUMN_ALIASES,
}

TANNER_CONTACT_COLUMN_LABELS = {
    "MAIL_AGENTE": {"mail_agente", "correo_agente"},
    "PHONO_AGENTE": {"phono_agente", "fono_agente", "telefono_agente", "telefono", "teléfono", "telefono1"},
}


def _normalize_agent_text(value: str) -> str:
    return " ".join((value or "").split()).strip()


def _series(df: pd.DataFrame, col: str) -> pd.Series:
    return pd.Series(df[col])


@dataclass
class MailTemplate:
    code: str
    label: str
    message_id: int
    institucion: str
    segmentoinstitucion: str
    mandante: str

    def display_label(self) -> str:
        return f"{self.label} (ID: {self.message_id})"


_DEFAULT_MAIL_TEMPLATE_OPTIONS: list[MailTemplate] = [
    MailTemplate(
        code="SC_TELEFONIA_DESCUENTO",
        label="Santander Consumer Telefonía - Descuento",
        message_id=95008,
        institucion="Santander Consumer",
        segmentoinstitucion="Santander Consumer",
        mandante="Santander Consumer Telefonía",
    ),
    MailTemplate(
        code="SC_TELEFONIA_MEDIOS_PAGO",
        label="Santander Consumer Telefonía - Medios de Pago",
        message_id=96706,
        institucion="Santander Consumer",
        segmentoinstitucion="Santander Consumer",
        mandante="Santander Consumer Telefonía",
    ),
    MailTemplate(
        code="ITAU_VENCIDA_MAIL",
        label="Itau Vencida - Mail",
        message_id=84824,
        institucion="BANCO ITAÚ",
        segmentoinstitucion="BANCO ITAÚ",
        mandante="Itau Vencida",
    ),
    MailTemplate(
        code="TANNER_MEDIOS_PAGO",
        label="Tanner - Medios de Pago",
        message_id=91869,
        institucion="TANNER SERVICIOS FINANCIEROS",
        segmentoinstitucion="TANNER",
        mandante="Tanner",
    ),
    MailTemplate(
        code="TANNER_CASTIGO",
        label="Tanner - Castigo",
        message_id=96830,
        institucion="TANNER SERVICIOS FINANCIEROS",
        segmentoinstitucion="TANNER",
        mandante="Tanner",
    ),
    MailTemplate(
        code="SCJ_COBRANZA",
        label="Santander Consumer Judicial - Cobranza",
        message_id=86257,
        institucion="SANTANDER CONSUMER",
        segmentoinstitucion="SANTANDER CONSUMER",
        mandante="Santander Consumer Judicial",
    ),
]


def _load_mail_templates_from_config() -> list[MailTemplate]:
    path = config_path("mail_templates.json")
    if not path.exists():
        return _DEFAULT_MAIL_TEMPLATE_OPTIONS

    try:
        raw_items = config_store.read_json("mail_templates.json")
        templates = [
            MailTemplate(
                code=str(item["code"]).strip(),
                label=str(item["label"]).strip(),
                message_id=int(item["message_id"]),
                institucion=str(item["institucion"]).strip(),
                segmentoinstitucion=str(item["segmentoinstitucion"]).strip(),
                mandante=str(item["mandante"]).strip(),
            )
            for item in raw_items
        ]
    except (OSError, ValueError, TypeError, KeyError, json.JSONDecodeError):
        return _DEFAULT_MAIL_TEMPLATE_OPTIONS
    return templates or _DEFAULT_MAIL_TEMPLATE_OPTIONS


MAIL_TEMPLATE_OPTIONS: list[MailTemplate] = _load_mail_templates_from_config()


def get_template_by_id(template_id: int) -> Optional[MailTemplate]:
    for template in MAIL_TEMPLATE_OPTIONS:
        if template.message_id == template_id:
            return template
    return None


def get_template(code: str) -> Optional[MailTemplate]:
    for template in MAIL_TEMPLATE_OPTIONS:
        if template.code == code:
            return template
    return None


def build_mail_template(df: pd.DataFrame, template_code: str, mandante: Optional[str] = None) -> pd.DataFrame:
    template = get_template(template_code)
    if not template:
        raise ValueError("Plantilla no soportada.")

    if template_code == "SC_TELEFONIA_DESCUENTO":
        return _build_sc_telefonia_descuento(df, template)
    if template_code == "SC_TELEFONIA_MEDIOS_PAGO":
        return _build_sc_telefonia_medios_pago(df, template)
    if template_code == "ITAU_VENCIDA_MAIL":
        from services.mail_itau_vencida import build_itau_vencida

        return build_itau_vencida(df, template, mandante)
    if template_code in ITAU_CASTIGO_SENDERS:
        return _build_itau_castigo(df, template)
    if template_code in BIT_TEMPLATE_CODES:
        return _build_bit_mail(df, template)
    if template_code in {"TANNER_MEDIOS_PAGO", "TANNER_CASTIGO"}:
        return _build_tanner_medios_pago(df, template, mandante)
    if template_code == "SCJ_COBRANZA":
        return _build_scj_cobranza(df, template, mandante)

    raise ValueError("Plantilla no implementada.")


def sample_mail_template(template_code: str) -> pd.DataFrame:
    template = get_template(template_code)
    if not template:
        raise ValueError("Plantilla no soportada.")

    if template_code == "SC_TELEFONIA_DESCUENTO":
        sample_df = pd.DataFrame({
            "NOMBRE_CLIENTE": ["CLIENTE PRUEBA UNO", "CLIENTE PRUEBA DOS"],
            "NRO_OPERACION": ["650080000001", "650080000002"],
            "MAIL": ["cliente1@example.com", "cliente2@example.com"],
        })
        return build_mail_template(sample_df, template_code, mandante="Santander Consumer Telefonía")
    if template_code == "SC_TELEFONIA_MEDIOS_PAGO":
        sample_df = pd.DataFrame({
            "RUT": ["11111111", "22222222"],
            "MAIL": ["cliente1@example.com", "cliente2@example.com"],
        })
        return build_mail_template(sample_df, template_code, mandante="Santander Consumer Telefonía")
    if template_code == "ITAU_VENCIDA_MAIL":
        sample_df = pd.DataFrame({
            "Oper": ["000000000000000002046954", "000000000000000002094727"],
            "RUT": ["13433958", "9561969"],
            "DV1": ["6", "K"],
            "Nombre": ["FELIPE EDUARDO RODAS KRAUSE", "ALBERTO AURELIO MENDEZ AMPUERO"],
            "MASIVIDAD": ["EMAIL", "EMAIL"],
            "EMAIL": ["felipe@example.com", "alberto@example.com"],
            "CARTERIZADO ABRIL": ["Jessica Carolina Diaz Mata", "Veronica Margarita Vega Bustos"],
        })
        return build_mail_template(sample_df, template_code, mandante="Itau Vencida")
    if template_code in ITAU_CASTIGO_SENDERS:
        sample_df = pd.DataFrame({
            "RUT": ["11111111-1", "11111111-1", "22222222-2"],
            "OPERACION": ["IC1", "IC2", "IC3"],
            "MAIL": ["primero@example.com", "duplicado-rut@example.com", "PRIMERO@EXAMPLE.COM"],
        })
        return build_mail_template(sample_df, template_code, mandante="Itau Castigo")
    if template_code in BIT_TEMPLATE_CODES:
        sample_df = pd.DataFrame({
            "RUT": ["11111111-1", "11111111-1", "22222222-2", "33333333-3"],
            "OPERACION": ["BIT1", "BIT2", "BIT3", "BIT4"],
            "CLIENTE": ["CLIENTE UNO", "CLIENTE DUP RUT", "CLIENTE DUP MAIL", "CLIENTE TRES"],
            "MAIL": ["primero.bit@example.com", "duplicado-rut@example.com", "PRIMERO.BIT@EXAMPLE.COM", "tercero.bit@example.com"],
        })
        return build_mail_template(sample_df, template_code, mandante="Banco Internacional")
    if template_code in {"TANNER_MEDIOS_PAGO", "TANNER_CASTIGO"}:
        sample_df = pd.DataFrame({
            "RUT+DV": ["11.111.111-1", "22.222.222-2"],
            "OPERACION": ["890123", "567890"],
            "dest_email": ["cliente1@example.com", "cliente2@example.com"],
            "NOMBRE_AGENTE": ["Juan Pérez", "Ana Díaz"],
            "MAIL_AGENTE": ["agente1@tanner.cl", "agente2@tanner.cl"],
            "PHONO_AGENTE": ["56912345678", "56987654321"],
            "name_from": ["Juan Pérez", "Ana Díaz"],
        })
        return build_mail_template(sample_df, template_code, mandante=None)
    if template_code == "SCJ_COBRANZA":
        sample_df = pd.DataFrame({
            "RUT": ["11.111.111-1", "22.222.222-2"],
            "NUM_OP": ["123456789", "987654321"],
            "dest_email": ["cliente1@example.com", "cliente2@example.com"],
            "NOMBRE_AGENTE": ["Ariel Silva", "Martina Parra"],
            "MAIL_AGENTE": ["asilva@phlegal.cl", "mparra@phlegal.cl"],
            "PHONO_AGENTE": ["56911111111", "56922222222"],
            "name_from": ["Ariel Silva", "Martina Parra"],
        })
        return build_mail_template(sample_df, template_code, mandante="Santander Consumer Judicial")

    raise ValueError("Plantilla no implementada para ejemplo.")


def _build_tanner_medios_pago(df: pd.DataFrame, template: MailTemplate, mandante: Optional[str]) -> pd.DataFrame:
    base = df.copy()
    base.columns = [str(col).strip() for col in base.columns]

    rut_col = _find_column(base, TANNER_REQUIRED_COLUMN_LABELS["RUT+DV"])
    dv_col = None
    if rut_col and "rut+dv" not in rut_col.lower().replace(" ", ""):
        dv_col = _find_column(base, {"dv", "digito", "dígito", "dv_rut"})
    op_col = _find_column(base, TANNER_REQUIRED_COLUMN_LABELS["OPERACION"])
    dest_col = _find_column(base, TANNER_REQUIRED_COLUMN_LABELS["dest_email"])
    name_from_col = _find_column(base, {"name_from", "nombre_remitente"})
    mail_agente_col = _find_column(base, TANNER_CONTACT_COLUMN_LABELS["MAIL_AGENTE"])
    nombre_agente_col = _find_column(base, TANNER_REQUIRED_COLUMN_LABELS["NOMBRE_AGENTE"])
    phono_col = _find_column(base, TANNER_CONTACT_COLUMN_LABELS["PHONO_AGENTE"])

    missing_required = [
        logical_name
        for logical_name, col in (
            ("RUT+DV", rut_col),
            ("OPERACION", op_col),
            ("dest_email", dest_col),
            ("NOMBRE_AGENTE", nombre_agente_col),
        )
        if col is None
    ]
    if missing_required:
        missing_details = {name: TANNER_REQUIRED_COLUMN_LABELS[name] for name in missing_required}
        raise ValueError(
            "Faltan columnas requeridas para la plantilla de Tanner: "
            + ", ".join(missing_required)
            + ". Encabezados aceptados: "
            + _format_expected_columns(missing_details)
        )

    require_fallback_columns = mandante is None
    if require_fallback_columns and (not mail_agente_col or not phono_col):
        missing_contact = [
            logical_name
            for logical_name, col in (
                ("MAIL_AGENTE", mail_agente_col),
                ("PHONO_AGENTE", phono_col),
            )
            if col is None
        ]
        missing_details = {name: TANNER_CONTACT_COLUMN_LABELS[name] for name in missing_contact}
        raise ValueError(
            "Faltan columnas de contacto del agente en el archivo de origen: "
            + ", ".join(missing_contact)
            + ". Encabezados aceptados: "
            + _format_expected_columns(missing_details)
        )

    base = dedupe_by_column_keep_first(base, rut_col).reset_index(drop=True)

    rut_series = base[rut_col].fillna("").astype(str).str.replace(r"\.0$", "", regex=True).str.strip()
    if rut_col and "rut+dv" not in rut_col.lower().replace(" ", "") and dv_col:
        dv_series = base[dv_col].fillna("").astype(str).str.replace(r"\.0$", "", regex=True).str.strip()
        rut_series = rut_series + "-" + dv_series

    op_series = base[op_col].fillna("").astype(str).str.replace(r"\.0$", "", regex=True).str.strip()

    month_name = SPANISH_MONTHS[datetime.now().month - 1]
    year_str = str(datetime.now().year)

    count = len(base)
    dest_values = base[dest_col].fillna("").astype(str).str.strip().tolist()
    nombre_agente_series = (
        base[nombre_agente_col]
        .fillna("")
        .astype(str)
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
    )
    name_from_base = (
        base[name_from_col]
        .fillna("")
        .astype(str)
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
        if name_from_col
        else nombre_agente_series
    )

    nombre_agente_values = []
    mail_agente_values = []
    phono_values = []
    mail_from_values = []
    name_from_values = []

    exec_cache: dict[tuple[str, str], Optional[ejecutivos_repo.Ejecutivo]] = {}

    def _get_ejecutivo(agente: str) -> Optional[ejecutivos_repo.Ejecutivo]:
        if not mandante or not agente:
            return None
        key = (mandante.lower(), agente.lower())
        if key not in exec_cache:
            exec_cache[key] = ejecutivos_repo.fetch_by_mandante_and_nombre(mandante, agente)
        return exec_cache[key]

    for idx, row in base.iterrows():
        agente_val = _normalize_agent_text(nombre_agente_series.iloc[idx])
        ejecutivo = _get_ejecutivo(agente_val)

        if ejecutivo:
            nombre_agente = ejecutivo.nombre_mostrar or agente_val
            mail_agente = ejecutivo.correo or ""
            phono = ejecutivo.telefono or ""
            mail_from = ejecutivo.reenviador or ejecutivo.correo or mail_agente or ""
            name_from = ejecutivo.nombre_mostrar or nombre_agente
        else:
            mail_agente = (
                str(row.get(mail_agente_col, "")).strip()
                if mail_agente_col
                else ""
            )
            phono = (
                str(row.get(phono_col, "")).strip()
                if phono_col
                else ""
            )
            nombre_agente = agente_val or _normalize_agent_text(name_from_base.iloc[idx])
            mail_from = mail_agente
            name_from = name_from_base.iloc[idx]

            if mandante and not mail_agente:
                raise ValueError(
                    f"No se encontró el ejecutivo '{agente_val}' para mandante {mandante} y el archivo no entrega correo del agente."
                )

        nombre_agente_values.append(_normalize_agent_text(nombre_agente))
        mail_agente_values.append(mail_agente)
        phono_values.append(phono)
        mail_from_values.append(mail_from)
        name_from_values.append(name_from or nombre_agente)

    data = {
        "INSTITUCIÓN": [template.institucion] * count,
        "SEGMENTOINSTITUCIÓN": [template.segmentoinstitucion] * count,
        "message_id": [template.message_id] * count,
        "RUT+DV": rut_series.tolist(),
        "OPERACION": op_series.tolist(),
        "dest_email": dest_values,
        "name_from": name_from_values,
        "mail_from": mail_from_values,
        "NOMBRE_AGENTE": nombre_agente_values,
        "MAIL_AGENTE": mail_agente_values,
        "PHONO_AGENTE": phono_values,
        "MES": [month_name] * count,
        "ANO": [year_str] * count,
    }

    output = pd.DataFrame(data)
    return output.reindex(columns=TEMPLATE_COLUMNS_TANNER)


ITAU_CASTIGO_SENDERS = {
    "ITAU_CASTIGO_SIN_DIRECCION_INGRID": {
        "name_from": "Ingrid Del Carmen Retamal Garrido",
        "mail_from": "iretamal@info.phoenixserviceinfo.cl",
        "CORREO": "iretamal@phoenixservice.cl",
    },
    "ITAU_CASTIGO_CON_DIRECCION_INGRID": {
        "name_from": "Ingrid Del Carmen Retamal Garrido",
        "mail_from": "iretamal@info.phoenixserviceinfo.cl",
        "CORREO": "iretamal@phoenixservice.cl",
    },
    "ITAU_CASTIGO_JL": {
        "name_from": "Jorge Francisco Lopez Cornejo",
        "mail_from": "jlopez@info.phoenixserviceinfo.cl",
        "CORREO": "jlopez@phoenixservice.cl",
    },
}

ITAU_CASTIGO_SEEDS = [
    {"RUT": "1-1", "OPERACION": "1234", "dest_email": "pipe5550@gmail.com"},
    {"RUT": "1-2", "OPERACION": "1234", "dest_email": "jriveros@phoenixservice.cl"},
]

BIT_TEMPLATE_CODES = {"BIT_CASTIGO", "BIT_VIGENTE"}
BIT_NAME_FROM = "Claudia Andrea Fuentes Fuentes"
BIT_MAIL_FROM = "cfuentes@info.phoenixserviceinfo.cl"
BIT_MAIL_AGENTE = "cfuentes@phoenixservice.cl"
BIT_SEEDS = [
    {"RUT": "1-1", "OPERACION": "1234", "CLIENTE": "PRB", "dest_email": "pipe5550@gmail.com"},
    {"RUT": "1-2", "OPERACION": "1234", "CLIENTE": "PRB", "dest_email": "cfuentes@phoenixservice.cl"},
]


def _build_itau_castigo(df: pd.DataFrame, template: MailTemplate) -> pd.DataFrame:
    base = df.copy()
    base.columns = [str(col).strip() for col in base.columns]

    rut_col = _find_column(base, RUT_COLUMN_ALIASES)
    oper_col = _find_column(base, OPERATION_COLUMN_ALIASES)
    dest_col = _find_column(base, EMAIL_COLUMN_ALIASES)

    missing = [
        name
        for name, col in (("RUT", rut_col), ("OPERACION", oper_col), ("dest_email", dest_col))
        if col is None
    ]
    if missing:
        raise ValueError("Faltan columnas requeridas para Itau Castigo: " + ", ".join(missing))

    base = dedupe_by_column_keep_first(base, rut_col).reset_index(drop=True)
    base = dedupe_by_column_keep_first_normalized(base, dest_col).reset_index(drop=True)

    sender = ITAU_CASTIGO_SENDERS.get(template.code, {})
    output = pd.DataFrame(
        {
            "INSTITUCIÓN": [template.institucion] * len(base),
            "SEGMENTOINSTITUCIÓN": [template.segmentoinstitucion] * len(base),
            "message_id": [template.message_id] * len(base),
            "PLANTILLA": ["CASTIGO"] * len(base),
            "RUT": base[rut_col].fillna("").astype(str).str.replace(r"\.0$", "", regex=True).str.strip().tolist(),
            "OPERACION": base[oper_col].fillna("").astype(str).str.replace(r"\.0$", "", regex=True).str.strip().tolist(),
            "dest_email": base[dest_col].fillna("").astype(str).str.strip().tolist(),
            "name_from": [sender.get("name_from", "")] * len(base),
            "mail_from": [sender.get("mail_from", "")] * len(base),
            "CORREO": [sender.get("CORREO", "")] * len(base),
        }
    )
    seed_rows = []
    for seed in ITAU_CASTIGO_SEEDS:
        seed_row = {
            "INSTITUCIÓN": template.institucion,
            "SEGMENTOINSTITUCIÓN": template.segmentoinstitucion,
            "message_id": template.message_id,
            "PLANTILLA": "CASTIGO",
            "RUT": seed["RUT"],
            "OPERACION": seed["OPERACION"],
            "dest_email": seed["dest_email"],
            "name_from": sender.get("name_from", ""),
            "mail_from": sender.get("mail_from", ""),
            "CORREO": sender.get("CORREO", ""),
        }
        seed_rows.append(seed_row)
    seed_df = pd.DataFrame(seed_rows).reindex(columns=TEMPLATE_COLUMNS_ITAU_CASTIGO)
    return pd.concat([seed_df, output.reindex(columns=TEMPLATE_COLUMNS_ITAU_CASTIGO)], ignore_index=True)


def _build_bit_mail(df: pd.DataFrame, template: MailTemplate) -> pd.DataFrame:
    base = df.copy()
    base.columns = [str(col).strip() for col in base.columns]

    rut_col = _find_column(base, RUT_COLUMN_ALIASES)
    oper_col = _find_column(base, OPERATION_COLUMN_ALIASES)
    cliente_col = _find_column(base, {"cliente", "nombre", "nombre_cliente", "contacto"})
    dest_col = _find_column(base, EMAIL_COLUMN_ALIASES)

    missing = [
        name
        for name, col in (("RUT", rut_col), ("OPERACION", oper_col), ("CLIENTE", cliente_col), ("dest_email", dest_col))
        if col is None
    ]
    if missing:
        raise ValueError("Faltan columnas requeridas para BIT: " + ", ".join(missing))

    base = dedupe_by_column_keep_first(base, rut_col).reset_index(drop=True)
    base = dedupe_by_column_keep_first_normalized(base, dest_col).reset_index(drop=True)

    output = pd.DataFrame(
        {
            "INSTITUCIÓN": [template.institucion] * len(base),
            "SEGMENTOINSTITUCIÓN": [template.segmentoinstitucion] * len(base),
            "message_id": [template.message_id] * len(base),
            "RUT": base[rut_col].fillna("").astype(str).str.replace(r"\.0$", "", regex=True).str.strip().tolist(),
            "OPERACION": base[oper_col].fillna("").astype(str).str.replace(r"\.0$", "", regex=True).str.strip().tolist(),
            "CLIENTE": base[cliente_col].fillna("").astype(str).str.strip().tolist(),
            "dest_email": base[dest_col].fillna("").astype(str).str.strip().tolist(),
            "name_from": [BIT_NAME_FROM] * len(base),
            "mail_from": [BIT_MAIL_FROM] * len(base),
            "MAIL_AGENTE": [BIT_MAIL_AGENTE] * len(base),
        }
    )
    seed_rows = []
    for seed in BIT_SEEDS:
        seed_rows.append(
            {
                "INSTITUCIÓN": template.institucion,
                "SEGMENTOINSTITUCIÓN": template.segmentoinstitucion,
                "message_id": template.message_id,
                "RUT": seed["RUT"],
                "OPERACION": seed["OPERACION"],
                "CLIENTE": seed["CLIENTE"],
                "dest_email": seed["dest_email"],
                "name_from": BIT_NAME_FROM,
                "mail_from": BIT_MAIL_FROM,
                "MAIL_AGENTE": BIT_MAIL_AGENTE,
            }
        )
    seed_df = pd.DataFrame(seed_rows).reindex(columns=TEMPLATE_COLUMNS_BIT)
    return pd.concat([seed_df, output.reindex(columns=TEMPLATE_COLUMNS_BIT)], ignore_index=True)


ITAU_SUPERVISOR = "Karen Avendaño"
ITAU_SUPERVISOR_MAIL = "kavendano@phoenixservice.cl"
ITAU_SUPERVISOR_PHONE = "442358500"
ITAU_CARTERA = "Vencida"
ITAU_CORREO_RECEPCION = "atencionclienteitauvigente@phoenixservice.cl"


def _ascii_fold(value: str) -> str:
    return unicodedata.normalize("NFKD", value or "").encode("ascii", "ignore").decode("ascii")


def _normalize_key(value: str) -> str:
    return (
        _ascii_fold(str(value or ""))
        .lower()
        .replace(" ", "")
        .replace("_", "")
        .replace("-", "")
    )


def _normalize_phone(value: Optional[str]) -> str:
    return re.sub(r"\D", "", str(value or ""))


def _prepare_itau_base(df: pd.DataFrame) -> pd.DataFrame:
    base = df.copy()
    base.columns = [str(col).strip() for col in base.columns]
    if _find_column(base, {"masividad"}):
        return base

    raw = df.copy()
    max_scan = min(len(raw), 15)
    header_idx = None
    required = {_normalize_key(x) for x in {"oper", "rut", "dv1", "nombre", "masividad", "email"}}
    for i in range(max_scan):
        vals = [str(v).strip() for v in raw.iloc[i].tolist()]
        keys = {_normalize_key(v) for v in vals if v}
        if required.issubset(keys):
            header_idx = i
            break
    if header_idx is None:
        return base

    headers = [str(v).strip() for v in raw.iloc[header_idx].tolist()]
    data = raw.iloc[header_idx + 1 :].copy()
    data.columns = headers
    data = data.dropna(how="all")
    data.columns = [str(col).strip() for col in data.columns]
    return data


def _resolve_itau_ejecutivo(mandante: str, agente: str) -> Optional[ejecutivos_repo.Ejecutivo]:
    raw = _normalize_agent_text(agente)
    if not raw:
        return None

    candidates = [
        raw,
        raw.lower(),
        raw.upper(),
        _ascii_fold(raw),
        _ascii_fold(raw).lower(),
        _ascii_fold(raw).upper(),
    ]
    seen: set[str] = set()
    unique = []
    for item in candidates:
        item = _normalize_agent_text(item)
        if item and item not in seen:
            seen.add(item)
            unique.append(item)

    for candidate in unique:
        found = ejecutivos_repo.fetch_by_mandante_and_nombre(mandante, candidate)
        if found:
            return found

    target = _ascii_fold(raw).lower().strip()
    if not target:
        return None
    best = None
    best_score = 0.0
    for item in ejecutivos_repo.list_ejecutivos(mandante=mandante, activos=True):
        options = [
            _ascii_fold(item.nombre_mostrar or "").lower().strip(),
            (item.nombre_clave or "").replace("_", " ").lower().strip(),
        ]
        for option in options:
            if not option:
                continue
            score = SequenceMatcher(None, target, option).ratio()
            if score > best_score:
                best_score = score
                best = item
    return best if best and best_score >= 0.92 else None


def _apply_current_itau_period(row: dict[str, str]) -> dict[str, str]:
    out = dict(row)
    out["MES_CURSO"] = SPANISH_MONTHS[datetime.now().month - 1].upper()
    out["ANO_CURSO"] = str(datetime.now().year)
    return out


def _load_itau_seed_rows_from_config() -> list[dict[str, str]]:
    path = config_path("mail_itau_vencida_seeds.json")
    if not path.exists():
        return []
    try:
        raw_items = config_store.read_json("mail_itau_vencida_seeds.json")
    except (OSError, json.JSONDecodeError):
        return []
    if not isinstance(raw_items, list):
        return []

    seeds = []
    for item in raw_items:
        if not isinstance(item, dict):
            continue
        row = {col: str(item.get(col) or "").strip() for col in TEMPLATE_COLUMNS_ITAU_VENCIDA}
        if any(row.values()):
            seeds.append(_apply_current_itau_period(row))
    return seeds


def _load_itau_seed_rows() -> list[dict[str, str]]:
    config_seeds = _load_itau_seed_rows_from_config()
    if config_seeds:
        return config_seeds

    path_candidates = [
        archive_path("mail_itau_vencida_excel_seed", "MAIL_VENCIDA_20260413.xlsx"),
        PROJECT_ROOT / "PLANTILLAS MAIL" / "ITAU VENCIDA" / "84824 ITAU VENCIDA MAIL" / "MAIL_VENCIDA_20260413.xlsx",
    ]
    path = next((candidate for candidate in path_candidates if candidate.exists()), None)
    if not path:
        return []

    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    target_map = {_normalize_key(col): col for col in TEMPLATE_COLUMNS_ITAU_VENCIDA}

    def _is_yellow(cell) -> bool:
        fill = cell.fill
        if not fill or fill.fill_type is None:
            return False
        fg = fill.fgColor
        if fg is None:
            return False
        if fg.type != "rgb":
            return False
        rgb = (fg.rgb or "").upper()
        return rgb.endswith("FFFF00") or rgb.endswith("FFD966")

    seeds: list[dict[str, str]] = []
    for row_idx in range(2, ws.max_row + 1):
        if not _is_yellow(ws.cell(row_idx, 1)):
            if seeds:
                break
            continue
        row_dict: dict[str, str] = {}
        for col_idx, header in enumerate(headers, start=1):
            if not header:
                continue
            mapped = target_map.get(_normalize_key(header))
            if not mapped:
                continue
            value = ws.cell(row_idx, col_idx).value
            row_dict[mapped] = str(value).strip() if value is not None else ""
        if any(str(v).strip() for v in row_dict.values()):
            seeds.append(_apply_current_itau_period(row_dict))

    return seeds


SCJ_PLANTILLA_VALUE = "CobranzaP"
SCJ_TELEFONO = "930609666"


def _build_scj_cobranza(df: pd.DataFrame, template: MailTemplate, mandante: Optional[str]) -> pd.DataFrame:
    base = df.copy()
    base.columns = [str(col).strip() for col in base.columns]

    rut_col = _find_column(base, RUT_COLUMN_ALIASES)
    dv_col = None
    if rut_col and "rut+dv" not in rut_col.lower().replace(" ", ""):
        dv_col = _find_column(base, {"dv", "digito", "dígito", "dv_rut"})
    op_col = _find_column(base, OPERATION_COLUMN_ALIASES)
    dest_col = _find_column(base, EMAIL_COLUMN_ALIASES)
    name_from_col = _find_column(base, {"name_from", "nombre_remitente"})
    mail_agente_col = _find_column(base, {"mail_agente", "correo_agente"})
    nombre_agente_col = _find_column(base, AGENTE_COLUMN_ALIASES)
    phono_col = _find_column(base, {"phono_agente", "fono_agente", "telefono_agente", "telefono"})

    if not (rut_col and op_col and dest_col):
        raise ValueError("Faltan columnas requeridas para la plantilla de Santander Consumer Judicial.")

    base = dedupe_by_column_keep_first(base, rut_col).reset_index(drop=True)

    rut_series = base[rut_col].fillna("").astype(str).str.replace(r"\.0$", "", regex=True).str.strip()
    if rut_col and "rut+dv" not in rut_col.lower().replace(" ", "") and dv_col:
        dv_series = base[dv_col].fillna("").astype(str).str.replace(r"\.0$", "", regex=True).str.strip()
        rut_series = rut_series + "-" + dv_series

    op_series = base[op_col].fillna("").astype(str).str.replace(r"\.0$", "", regex=True).str.strip()
    dest_values = base[dest_col].fillna("").astype(str).str.strip().tolist()

    nombre_agente_series = (
        base[nombre_agente_col]
        .fillna("")
        .astype(str)
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
        if nombre_agente_col else pd.Series([""] * len(base), index=base.index)
    )
    name_from_base = (
        base[name_from_col]
        .fillna("")
        .astype(str)
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
        if name_from_col else nombre_agente_series
    )

    nombre_agente_values = []
    mail_agente_values = []
    phono_values = []
    mail_from_values = []
    name_from_values = []

    exec_cache: dict[tuple[str, str], Optional[ejecutivos_repo.Ejecutivo]] = {}

    def _get_ejecutivo(agente: str) -> Optional[ejecutivos_repo.Ejecutivo]:
        if not mandante or not agente:
            return None
        key = (mandante.lower(), agente.lower())
        if key not in exec_cache:
            exec_cache[key] = ejecutivos_repo.fetch_by_mandante_and_nombre(mandante, agente)
        return exec_cache[key]

    for idx, row in base.iterrows():
        agente_val = _normalize_agent_text(nombre_agente_series.iloc[idx])
        ejecutivo = _get_ejecutivo(agente_val)

        if ejecutivo:
            nombre_agente = ejecutivo.nombre_mostrar or agente_val or name_from_base.iloc[idx]
            mail_agente = ejecutivo.correo or ""
            phono = ejecutivo.telefono or ""
            mail_from = ejecutivo.reenviador or ejecutivo.correo or mail_agente or ""
            name_from = ejecutivo.nombre_mostrar or nombre_agente
        else:
            mail_agente = (
                str(row.get(mail_agente_col, "")).strip()
                if mail_agente_col else ""
            )
            phono = (
                str(row.get(phono_col, "")).strip()
                if phono_col else ""
            )
            nombre_agente = agente_val or _normalize_agent_text(name_from_base.iloc[idx])
            mail_from = mail_agente
            name_from = name_from_base.iloc[idx]

            if mandante and not mail_agente:
                raise ValueError(
                    f"No se encontró el ejecutivo '{agente_val}' para mandante {mandante} y el archivo no entrega correo del agente."
                )

        nombre_agente_values.append(_normalize_agent_text(nombre_agente))
        mail_agente_values.append(mail_agente)
        phono_values.append(phono)
        mail_from_values.append(mail_from or mail_agente)
        name_from_values.append(name_from or nombre_agente)

    data = {
        "INSTITUCIÓN": [template.institucion] * len(base),
        "SEGMENTOINSTITUCIÓN": [template.segmentoinstitucion] * len(base),
        "message_id": [template.message_id] * len(base),
        "PLANTILLA": [SCJ_PLANTILLA_VALUE] * len(base),
        "RUT": rut_series.tolist(),
        "NUM_OP": op_series.tolist(),
        "name_from": name_from_values,
        "dest_email": dest_values,
        "mail_from": mail_from_values,
        "MAIL_AGENTE": mail_agente_values,
        "PHONO_AGENTE": phono_values,
        "telefono": [SCJ_TELEFONO] * len(base),
        "NOMBRE_AGENTE": nombre_agente_values,
    }

    output = pd.DataFrame(data)
    return output.reindex(columns=TEMPLATE_COLUMNS_SCJ)


SC_TELEFONIA_PLANTILLA = "TEMPRANA"
SC_TELEFONIA_NAME_FROM = "Atencion Cliente Consumer"
SC_TELEFONIA_MAIL_FROM = "tfernandez@info.phoenixserviceinfo.cl"
SC_TELEFONIA_CORREO = "tfernandez@phoenixservice.cl"
SEED_EMAIL = "pipe5550@gmail.com"

SC_TELEFONIA_MP_PLANTILLA = "TEMPRANA"
SC_TELEFONIA_MP_NAME_FROM = "Atencion Cliente Consumer"
SC_TELEFONIA_MP_MAIL_FROM = "atencionclientes@estandar.phoenixserviceinfo.cl"
SC_TELEFONIA_MP_CORREO = "tfernandez@phoenixservice.cl"


def _load_sc_telefonia_seed_rows() -> list[dict[str, str]]:
    return [{
        "INSTITUCIÓN": "Santander Consumer",
        "SEGMENTOINSTITUCIÓN": "Santander Consumer",
        "message_id": "95008",
        "PLANTILLA": SC_TELEFONIA_PLANTILLA,
        "CLIENTE": "PRB",
        "NRO_OPERACION": "PRB",
        "dest_email": SEED_EMAIL,
        "name_from": SC_TELEFONIA_NAME_FROM,
        "mail_from": SC_TELEFONIA_MAIL_FROM,
        "CORREO": SC_TELEFONIA_CORREO,
    }]


def _build_sc_telefonia_descuento(df: pd.DataFrame, template: MailTemplate) -> pd.DataFrame:
    base = df.copy()
    base.columns = [str(col).strip() for col in base.columns]

    cliente_col = _find_column(base, {"nombre_cliente", "cliente", "nombre"})
    oper_col = _find_column(base, OPERATION_COLUMN_ALIASES)
    email_col = _find_column(base, EMAIL_COLUMN_ALIASES)

    if not (cliente_col and oper_col and email_col):
        raise ValueError("Faltan columnas requeridas para Santander Consumer Telefonía (cliente, operacion, mail).")

    cliente = base[cliente_col].fillna("").astype(str).str.strip()
    oper = base[oper_col].fillna("").astype(str).str.replace(r"\.0$", "", regex=True).str.strip()
    dest_email = base[email_col].fillna("").astype(str).str.strip()

    output = pd.DataFrame({
        "INSTITUCIÓN": [template.institucion] * len(base),
        "SEGMENTOINSTITUCIÓN": [template.segmentoinstitucion] * len(base),
        "message_id": [template.message_id] * len(base),
        "PLANTILLA": [SC_TELEFONIA_PLANTILLA] * len(base),
        "CLIENTE": cliente.tolist(),
        "NRO_OPERACION": oper.tolist(),
        "dest_email": dest_email.tolist(),
        "name_from": [SC_TELEFONIA_NAME_FROM] * len(base),
        "mail_from": [SC_TELEFONIA_MAIL_FROM] * len(base),
        "CORREO": [SC_TELEFONIA_CORREO] * len(base),
    }).reindex(columns=TEMPLATE_COLUMNS_SC_TELEFONIA)

    seeds = _load_sc_telefonia_seed_rows()
    if not seeds:
        return output
    seed_df = pd.DataFrame(seeds).reindex(columns=TEMPLATE_COLUMNS_SC_TELEFONIA).fillna("")
    return pd.concat([seed_df, output], ignore_index=True)


def _build_sc_telefonia_medios_pago(df: pd.DataFrame, template: MailTemplate) -> pd.DataFrame:
    base = df.copy()
    base.columns = [str(col).strip() for col in base.columns]

    rut_col = _find_column(base, RUT_COLUMN_ALIASES)
    email_col = _find_column(base, EMAIL_COLUMN_ALIASES)

    if not (rut_col and email_col):
        raise ValueError("Faltan columnas requeridas para Medios de Pago Telefonía (RUT y MAIL).")

    base = dedupe_by_column_keep_first(base, rut_col).reset_index(drop=True)

    rut = base[rut_col].fillna("").astype(str).str.replace(r"\.0$", "", regex=True).str.strip()
    dest_email = base[email_col].fillna("").astype(str).str.strip()

    output = pd.DataFrame({
        "INSTITUCIÓN": [template.institucion] * len(base),
        "SEGMENTOINSTITUCIÓN": [template.segmentoinstitucion] * len(base),
        "message_id": [template.message_id] * len(base),
        "PLANTILLA": [SC_TELEFONIA_MP_PLANTILLA] * len(base),
        "RUT": rut.tolist(),
        "dest_email": dest_email.tolist(),
        "name_from": [SC_TELEFONIA_MP_NAME_FROM] * len(base),
        "mail_from": [SC_TELEFONIA_MP_MAIL_FROM] * len(base),
        "CORREO": [SC_TELEFONIA_MP_CORREO] * len(base),
    })
    output = output.reindex(columns=TEMPLATE_COLUMNS_SC_TELEFONIA_MEDIOS_PAGO)
    seed = pd.DataFrame([
        {
            "INSTITUCIÓN": "Santander Consumer",
            "SEGMENTOINSTITUCIÓN": "Santander Consumer",
            "message_id": "96706",
            "PLANTILLA": SC_TELEFONIA_MP_PLANTILLA,
            "RUT": "PRB",
            "dest_email": SEED_EMAIL,
            "name_from": SC_TELEFONIA_MP_NAME_FROM,
            "mail_from": SC_TELEFONIA_MP_MAIL_FROM,
            "CORREO": SC_TELEFONIA_MP_CORREO,
        }
    ]).reindex(columns=TEMPLATE_COLUMNS_SC_TELEFONIA_MEDIOS_PAGO)
    return pd.concat([seed, output], ignore_index=True)


def _find_column(df: pd.DataFrame, candidates: set[str]) -> Optional[str]:
    def norm(value: str) -> str:
        return (
            value.lower()
            .replace(" ", "")
            .replace("_", "")
            .replace("-", "")
        )

    lowered = {norm(col): col for col in df.columns}
    for candidate in candidates:
        key = norm(candidate)
        if key in lowered:
            return lowered[key]
    return None


def _format_expected_columns(field_map: dict[str, set[str]]) -> str:
    parts = []
    for logical_name, aliases in field_map.items():
        alias_list = ", ".join(sorted(aliases))
        parts.append(f"{logical_name} ({alias_list})")
    return "; ".join(parts)


